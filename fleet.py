from openpyxl import load_workbook, Workbook

# Load fleet data
wb = load_workbook('fleet_data.xlsx')
ws = wb.active

# Create report workbook
report_wb = Workbook()
report_ws = report_wb.active
report_ws.title = "Utilization Report"

# Write headers
report_ws.append(["Vehicle ID", "Branch", "Utilization %"])

# Read data, skip rows with any empty cells (None)
for row in ws.iter_rows(min_row=2, values_only=True):
    if any(cell is None for cell in row[:4]):  # skip if any cell empty
        continue
    vehicle_id, branch, total_days, days_rented = row[:4]
    utilization = (days_rented / total_days) * 100 if total_days else 0
    report_ws.append([vehicle_id, branch, round(utilization, 2)])

# Save report
report_wb.save('utilization_report.xlsx')
print("Report created!")